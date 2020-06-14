"""
This script automatically updates the MsLog file. 

A GUI needs to be added so that the user
can select the:
1) UpdateMsLog file
2) MsLogFor2018-2020 file

"""


"""
##STEPS INVOLVED IN MODIFYING THE MS_UPDATE FILE:

Step 1 -- DONE
Delete all MS submissions earlier than 2018

Step 2 -- DONE
Rename headers appropriately:
-remove "Title" row completely
-rename "Latest decision"	 	header to 	"Status"
-rename "# Reviewers Invited 	header to 	"Peer review"
-rename "First Decision Date" 	header to 	"1st decision"
-rename "Latest Decision Date" 	header to 	"Final decision"

Step 3 -- DONE
Replace "blank" cells in "Status" column with the cell data from the "Manuscript Status" column

Step 4  -- DONE
Filter for "draft", "complete checklist", and "pending payment" from the "Status" column and delete all MS submissions with these status labels

Step 5 -- DONE
Filter the data in the "Status" column and rename the items such that the following exist only: "Editorial Assessment", "Rejected", "Withdrawn", "Accepted" 

Step 6 -- DONE
Convert data in the column "Peer review" to either "Yes" or "No"



##STEPS INVOLVED IN UPDATING THE MS_LOG FILE:
 
Step 1  -- DONE
Remove everything from the old LogUpdate sheet

Step 2  -- DONE
Add the new LogUpdate values from the modified LogUpdate file

Step 3 -- DONE
Apply match formulas to the appropriate sections of the 2018-2020 sheet

Step 4 -- DONE
If publication date or doi values are in the MsLog, then change the status to "Published"

Step 5 -- DONE
recreate all of the data validations
"""



#Imports
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

#File names
UpdateMSLog_File = "UpdateMSlog_2019.12.03.xlsx"
UpdateMSLog_File_New_Temp = "tmp.xlsx"
MSLogFor2018_2020_File = "MS log for 2018-2020_fixed.xlsx"
MSLogFor2018_2020_UPDATED_FILE="MS log for 2018-2020_updated.xlsx"

#Load the updateLog file
print("Loading \"%s\"" % (UpdateMSLog_File))
wb2 = load_workbook(filename=UpdateMSLog_File, data_only=True)
ws2 = wb2.active

#Clean the headers in the updateLog file
try:
	ws2.unmerge_cells('A1:G1')
except:
	pass

print("\t-Renaming headers")
ws2.delete_rows(1)
ws2["D1"].value = "Status" 
ws2["E1"].value = "Peer review"
ws2["F1"].value = "1st decision"
ws2["G1"].value = "Final decision"

#Remove all entries earlier than 2018 and any blank ones from the updateLog file
print("\t-Removing all rows prior to 2018 (and blank entries) [takes about 30 seconds]")
i=1
del_rows = []
unwanted_entries = ["2017", "2016", "2015"]
for row in ws2.iter_rows(min_col=2, max_col=2, min_row=2):
	i += 1
	for cell in row:
		str_cell = str(cell.value)

		for x in unwanted_entries:
			if str_cell.startswith(x):
				del_rows.append(i)

		if cell.value is None:
			del_rows.append(i)

for r in reversed(del_rows):
	ws2.delete_rows(r)

#If "blank" Status values, then rename as Manuscript Status values in updateLog file
print("\t-Renaming empty Status column values with Manuscript Status values")
i=1
row_num = []
for row in ws2.iter_rows(min_col=4, max_col=4, min_row=2):
	for cell in row:
		i += 1
		if cell.value is None:
			row_num.append(i)

for r in row_num:
	cell_with_value = "C" + str(r)
	cell_that_is_blank = "D" + str(r)
	ws2[cell_that_is_blank].value = ws2[cell_with_value].value


#Remove all entries that have "Pending Payment Agreements", "Complete Checklist" or "Draft" values in the Manuscript Status column in the updateLog file
print("\t-Removing rows with Pending Payment, etc., values in the Manuscript Status column")
i=1
del_rows = []
for row in ws2.iter_rows(min_col=3, max_col=3, min_row=2):
	i += 1
	for cell in row:
		str_cell = str(cell.value)
		if str_cell.startswith("Pending Payment Agreements"):
			del_rows.append(i)

		if str_cell.startswith("Complete Checklist"):
			del_rows.append(i)

		if str_cell.startswith("Draft"):
			del_rows.append(i)

for r in reversed(del_rows):
	ws2.delete_rows(r)


#Rename all "status" column values to "Editorial Assessment", "Rejected", "Withdrawn", or "Accepted" in the updateLog file
print("\t-Renaming all Status entries to: Editorial Assessment, Rejected, Withdrawn, or Accepted")
accept = ["Accept"]
editorial_assessment = ["Reviewer", "Invite", "Make", "Revision" ]
withdrawn = [""]
rejected = ["Reject"]

for row in ws2.iter_rows(min_col=4, max_col=4, min_row=2):
	for cell in row:
		for x in rejected:
			if x in cell.value:
				cell.value = "Rejected"
		for x in editorial_assessment:
			if x in cell.value:
				cell.value = "Editorial Assessment"
		for x in accept:
			if x in cell.value:
				cell.value = "Accepted"	

#Rename all "Peer review" column values to "Yes" or "No" in the updateLog file
print("\t-Renaming Peer review values to: Yes or No")
for row in ws2.iter_rows(min_col=5, max_col=5, min_row=2):
	for cell in row:
		
		try:
			cell.value = int(cell.value)
		except:
			pass
		
		if cell.value > 0:
			cell.value = "Yes"
		else:
			cell.value = "No"

#Save the modifications made to the updateLog file as a new temporary file
print("\t-Saving all modifications to: \"%s\"" % (UpdateMSLog_File_New_Temp))
wb2.save(filename=UpdateMSLog_File_New_Temp)

#Load the MSLogfor2018-2020 file
print("\nLoading: \"%s\"" % (MSLogFor2018_2020_File))
wb1 = load_workbook(filename=MSLogFor2018_2020_File)

#Identify the sheets within the MSLogfor2018-2020 file
print("\t-Sheets found in the MSLogFor2018-2020 file: " + ', '.join(wb1.sheetnames))
ms_update_sheet = wb1["LogUpdate"]
validation_sheet = wb1["Dropdowns and documentation"]
ms_log_sheet = wb1["2018-2020"]
countries_sheet = wb1["Countries"]

#Clear all cell values from the LogUpdate sheet in the MSLogfor2018-2020 file
print("\t-Clearing content from the LogUpdate sheet")
for row in ms_update_sheet:
	for cell in row:
		cell.value = None

#Load the temporary updateLog file created:
print("\t-Loading temp file: %s" % (UpdateMSLog_File_New_Temp))
wb2 = load_workbook(filename=UpdateMSLog_File_New_Temp)
ws2 = wb2.active

mr = ws2.max_row
mc = ws2.max_column

#Copy the modified contents of the temporary updateLog file to the LogUpdate sheet of the MSLogfor2018-2020 file
print("\t-Copying new LogUpdate values from %s to the LogUpdate sheet of the MSLogfor2018-2020 file" % (UpdateMSLog_File_New_Temp))
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		c = ws2.cell(row = i, column = j)

		ms_update_sheet.cell(row = i, column = j).value = c.value



#If there are matches between the MS-IDs of the two files, then update the columns of the MSLogfor2018-2020 2018-2020 sheet to match the values within the columns of the updated LogUpdate sheet
print("\t-Performing MATCH function and updating values for Status, Peer Review, 1st Decision, Final Decision in the MSLogfor2018-2020 file")
list_of_MS_IDs = []
for row in ms_update_sheet.iter_rows(min_col=1, max_col=1, min_row=2):
	for cell in row:
		list_of_MS_IDs.append(cell.value)

row_num = 1
for row in ms_log_sheet.iter_rows(min_col=4, max_col=4, min_row=2):
	row_num += 1
	for cell in row:
		for x in list_of_MS_IDs:
			if x == cell.value:
				y = list_of_MS_IDs.index(x) + 2
				
				stat_cell = ms_update_sheet.cell(row = y, column = 4)
				peerRev_cell = ms_update_sheet.cell(row= y, column= 5)
				firstDec_cell = ms_update_sheet.cell(row= y, column= 6)
				lastDec_cell = ms_update_sheet.cell(row= y, column= 7)
				
				ms_log_sheet.cell(row = row_num, column = 10).value = stat_cell.value
				ms_log_sheet.cell(row = row_num, column = 11).value = peerRev_cell.value
				ms_log_sheet.cell(row = row_num, column = 12).value = firstDec_cell.value
				ms_log_sheet.cell(row = row_num, column = 13).value = lastDec_cell.value


#Update 2018-2020 Status column if there is anything in the "DOI" column so that the status value is "Published" instead of "Accepted"
print("\t-Updating MsLog Status values to \"Published\" if a DOI value has been added")
row_num = 1
for row in ms_log_sheet.iter_rows(min_col = 10, max_col = 10, min_row = 2):
	row_num += 1
	for cell in row:
		doi_status = ms_log_sheet.cell(row = row_num, column = 15).value
		if doi_status is not None:
			ms_log_sheet.cell(row = row_num, column = 10).value = "Published"
		else:
			pass



#Adding data validation to the 2018-2020 sheet of the MSLogfor2018-2020 file
print("\t-Adding Data Validation lists to the various columns of the 2018-2020 Sheet in the MSLogfor2018-2020 file")
dv_Status = DataValidation(type="list", formula1="='Dropdowns and documentation'!$A$3:$A$9", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_Status)
dv_Status.add('J2:J1048576')

dv_ArticleCategory = DataValidation(type="list", formula1="='Dropdowns and documentation'!$B$3:$B$12", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_ArticleCategory)
dv_ArticleCategory.add('G2:G1048576')

dv_Disciple = DataValidation(type="list", formula1="='Dropdowns and documentation'!$C$3:$C$9", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_Disciple)
dv_Disciple.add('H2:H1048576')

dv_PartOfASupplement = DataValidation(type="list", formula1="='Dropdowns and documentation'!$D$3:$D$4", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_PartOfASupplement)
dv_PartOfASupplement.add('I2:I1048576')

dv_GenderInclusion = DataValidation(type="list", formula1="='Dropdowns and documentation'!$E$3:$E$9", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_GenderInclusion)
dv_GenderInclusion.add('Z2:Z1048576')

dv_CountryClassification = DataValidation(type="list", formula1="='Dropdowns and documentation'!$F$3:$F$5", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_CountryClassification)
dv_CountryClassification.add('W2:W1048576')

dv_AuthorCountry = DataValidation(type="list", formula1="='Dropdowns and documentation'!$G$3:$G$4", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_AuthorCountry)
dv_AuthorCountry.add('T2:T1048576')
dv_AuthorCountry.add('U2:U1048576')
dv_AuthorCountry.add('V2:V1048576')

dv_Countries = DataValidation(type="list", formula1="='Countries'!$A$2:$A$219", showDropDown = 0, allow_blank = 1)
ms_log_sheet.add_data_validation(dv_Countries)
dv_Countries.add('P2:P1048576')
dv_Countries.add('Q2:Q1048576')
dv_Countries.add('R2:R1048576')

#Save the modifications made to the MSLogfor2018-2020 file to a new file
print("\t-Saving as: \"%s\"" % (MSLogFor2018_2020_UPDATED_FILE))
wb1.save(filename=MSLogFor2018_2020_UPDATED_FILE)