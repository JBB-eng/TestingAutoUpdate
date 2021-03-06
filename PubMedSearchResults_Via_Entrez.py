#search pubmed

import urllib.request, urllib.parse
import re, cloudscraper, os, openpyxl

from Bio import Entrez
from Bio import Medline
from openpyxl import load_workbook
from bs4 import BeautifulSoup

jias_bool = 1 #set to True if looking for JIAS publications
citation_bool = 1 #set to True to include citation amount for JIAS publications **SIGNIFICANTLY INCREASES PROCESSING TIME***

#SEARCH STRING
search_string = r'("J Int AIDS Soc"[jour]) AND ("2018"[Date - Publication] : "3000"[Date - Publication])) AND ((((Stigma[Title/Abstract]) OR Discrimination[Title/Abstract]) OR Criminalization[Title/Abstract]) OR "Human Rights"[Title/Abstract])'
search_string = r'("J Int AIDS Soc"[jour]) AND ("2019/01/01"[Date - Publication] : "2020/01/01"[Date - Publication])'
search_string = r'("J Int AIDS Soc"[Jour]) AND ("2015"[Date - Publication] : "3000"[Date - Publication]) AND Uganda[Title/Abstract])'
search_string = r'((J Int AIDS Soc[Jour]) AND ("2010"[Date - Publication] : "3000"[Date - Publication])) AND "South Africa"[Title/Abstract]'
search_string = r'((J Int AIDS Soc[Jour]) AND ("2012"[Date - Publication] : "3000"[Date - Publication])) AND "South Africa"[Title/Abstract]'

def GetDownloadPath():
	"""Returns the default downloads path for linux or windows"""
	if os.name == 'nt':
		import winreg
		sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
		downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
		with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
			location = winreg.QueryValueEx(key, downloads_guid)[0]
		return location
	else:
		return os.path.join(os.path.expanduser('~'), 'downloads')


def Scrap_For_TOC(url):
	"""scrapes the html for the ToC information"""
	issue = ""
	publication_type = ""
	publication_info = []
	scraper = cloudscraper.create_scraper()
	web_text = scraper.get(url).text
	
	soup = BeautifulSoup(web_text, features="lxml")
	temp_soup = soup.get_text().replace('\n\n', '')
	temp_soup = ''.join(temp_soup)
	temp_soup = temp_soup.split('\n')
	parsed_soup = []
	start_stop_parsing = 0
	start_stop_pub_info = 0
	#x=0
	for line in temp_soup:#soup.get_text():
		#print(line)
		if 'Facebook pageRSS FeedsMost recent' in line:
			start_stop_parsing = 1
			#print("start_stop = 1")
		elif 'ToolsSubmit an Article' in line:
			start_stop_parsing = 0
			#print("start_stop = 0")
		if start_stop_parsing == 1:
			parsed_soup.append(str(line))
			#print(line)
		if 'Select / Deselect allExport Citation(s)Export' in line:
			issue = line
		if 'Open Access' in line:
			start_stop_pub_info = 1
		elif 'Full text' in line:
			start_stop_pub_info = 0
		if start_stop_pub_info == 1 and start_stop_parsing == 1:
			publication_info.append(line)
			#x += 1
	print(issue)
	print(publication_info)


def GetCitationNumber(url):
	"""Gets number of citations for JIAS publications from Wiley's website"""
	scraper = cloudscraper.create_scraper()
	web_text = scraper.get(url).text
	
	citation_text = 'Citations: <a href="#citedby-section">'

	for line in web_text.splitlines():
		if citation_text in line:
			index = line.find(citation_text)
			temp = line[index:]
			temp = re.sub(citation_text, '', temp)
			temp = re.sub('</a></span></div>', '', temp)
			citation_number = temp
			break
		else:
			citation_number = 0

	return citation_number

#search Pubmed...
Entrez.email = "hello_world@example.com"  
handle = Entrez.egquery(term=search_string)
record = Entrez.read(handle)
count = 0
for row in record["eGQueryResult"]:
		if row["DbName"]=="pubmed":
			print("Number of articles found with requested search parameters:", row["Count"])
			count = row["Count"]

handle = Entrez.esearch(db="pubmed", term=search_string, retmax=count)
record = Entrez.read(handle)
handle.close()
idlist = record["IdList"]

handle = Entrez.efetch(db="pubmed", id=idlist, rettype="medline", retmode="text")
records = Medline.parse(handle)

records = list(records)
#print(records[142].get("SO", "?"))
#print(records[142].get("DP", "?"))

"""
#print results (without citation)...
x=1
for record in records:
	print("(" + str(x) + ")")
	print("Title:", record.get("TI", "?"))
	print("Authors: ", ", ".join(record.get("AU", "?")))
	print("Pub Date: " + record.get("DP", "?")[5:] + " " + record.get("DP", "?")[:-4])
	print("Journal:", record.get("JT", "?"))
	print("DOI:", record.get("LID", "?")[:-6])
	if jias_bool:
		print("Wiley Link: " + "https://onlinelibrary.wiley.com/doi/full/" + record.get("LID", "?")[:-6])
	print("\n")
	x += 1
"""


#add results to excel folder (including citations)...
x=1
filepath = GetDownloadPath() + "\\test_keyword_search.xlsx"
try:
	book = load_workbook(filepath)
except:
	print("creating excel file....")
	wb = openpyxl.Workbook()
	wb.save(filepath)
	book = load_workbook(filepath)
	sheet = book.active
	sheet['A1'] = "No."
	sheet['B1'] = "Title"
	sheet['C1'] = "Authors"
	sheet['D1'] = "Publication Date"
	sheet['E1'] = "DOI"
	sheet['F1'] = "Wiley Link"
	sheet['G1'] = "Citations"

ws = book.worksheets[0]
for record in records:
	for cell in ws["A"]:
		if cell.value is None:
			emptyRow = cell.row
			break
	else:
		emptyRow = cell.row + 1

	title = record.get("TI", "?")
	authors = ", ".join(record.get("AU", "?"))
	if authors == "?":
		authors = ", ".join(record.get("IR", "?"))
	pub_date = record.get("DP", "?")[5:] + " " + record.get("DP", "?")[:-4]
	#pub_date = re.sub(r'^.*?.', '', record.get("SO","?"))
	pub_date = re.sub(r'^.*?Soc. ', '', record.get("SO","?"))
	pub_date = record.get("DP", "?")
	if len(pub_date) < 9:
		pub_date = re.sub(r'^.*?Soc. ', '', record.get("SO","?"))
		pub_date = record.get("SO","?")
		pub_date = record.get("SO","?").replace(record.get("TA", "?"), '')
		tmp1 = pub_date.split(". ", 1)
		tmp2 = tmp1[1].split(";", 1)
		pub_date = tmp2[0]

	#pub_date = record.get("SO","?")
	journal_name = record.get("JT", "?")
	doi = record.get("LID", "?")[:-6]
	wiley_link = "http://onlinelibrary.wiley.com/doi/full/" + record.get("LID", "?")[:-6]

	if citation_bool:
		excel_data = [str(x), title, authors, pub_date, doi, wiley_link, GetCitationNumber(wiley_link)]
	else:
		excel_data = [str(x), title, authors, pub_date, doi, wiley_link, "skipped"]
	
	x += 1

	for row in excel_data:
		ws.append(excel_data)
		break

	book.save(filepath)


#Scrap_For_TOC("https://onlinelibrary.wiley.com/toc/17582652/2020/23/4")