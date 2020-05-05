"""
pub med search tool for JIAS with user-friendly GUI 
"""

__author__		= 'Jacob Bursavich'
__copyright__	= 'Copyright (C) 2020, Jacob Bursavich'
__credits__		= ['Jacob Bursavich']
__license__		= 'The MIT License (MIT)'
__maintainer__	= 'Jacob Bursavich'
__email__		= 'jbursavich@gmail.com'
__status__		= 'Beta'

__AppName__		= 'JIAS PubMed Search Tool'
__version__		= '0.1'


#imports
import tkinter as tk
from tkinter import *
import urllib.request, urllib.parse
import re, cloudscraper, os, openpyxl
from Bio import Entrez
from Bio import Medline
from openpyxl import load_workbook

global jias_bool, citation_bool
jias_bool = 1 #set to True if looking for JIAS publications
citation_bool = 1 #set to True to include citation amount for JIAS publications **SIGNIFICANTLY INCREASES PROCESSING TIME***

#SEARCH STRING
search_string = r'("J Int AIDS Soc"[jour]) AND ("2018"[Date - Publication] : "3000"[Date - Publication])) AND ((((Stigma[Title/Abstract]) OR Discrimination[Title/Abstract]) OR Criminalization[Title/Abstract]) OR "Human Rights"[Title/Abstract])'

def GetDownloadPath():
	"""Returns the default downloads path for linux or windows"""
	if os.name == 'nt':
		import winreg
		sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
		downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
		with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
			location = winreg.QueryValueEx(key, downloads_guid)[0]
		return location
	else:
		return os.path.join(os.path.expanduser('~'), 'downloads')


def GetCitationNumber(url):
	"""Gets number of citations for JIAS publications from Wiley's website"""
	scraper = cloudscraper.create_scraper()
	web_text = scraper.get(url).text
	
	citation_text = 'Citations: <a href="#citedby-section">'

	for line in web_text.splitlines():
		if citation_text in line:
			index = line.find(citation_text)
			temp = line[index:]
			temp = re.sub(citation_text, '', temp)
			temp = re.sub('</a></span></div>', '', temp)
			citation_number = temp
			break
		else:
			citation_number = 0

	return citation_number



def PubMedSearch(search_keyword):
	#search Pubmed...
	Entrez.email = "hello_world@example.com"  
	handle = Entrez.egquery(term=search_keyword)
	record = Entrez.read(handle)
	count = 0
	for row in record["eGQueryResult"]:
			if row["DbName"]=="pubmed":
				print("Number of articles found with requested search parameters:", row["Count"])
				count = row["Count"]

	handle = Entrez.esearch(db="pubmed", term=search_keyword, retmax=count)
	record = Entrez.read(handle)
	handle.close()
	idlist = record["IdList"]

	handle = Entrez.efetch(db="pubmed", id=idlist, rettype="medline", retmode="text")
	records = Medline.parse(handle)

	records = list(records)
	handle.close()

	#print results (without citation)...
	x=1
	for record in records:
		print("(" + str(x) + ")")
		print("Title:", record.get("TI", "?"))
		print("Authors: ", ", ".join(record.get("AU", "?")))
		print("Pub Date: " + record.get("DP", "?")[5:] + " " + record.get("DP", "?")[:-4])
		print("Journal:", record.get("JT", "?"))
		print("DOI:", record.get("LID", "?")[:-6])
		if jias_bool:
			print("Wiley Link: " + "https://onlinelibrary.wiley.com/doi/full/" + record.get("LID", "?")[:-6])
		print("\n")
		x += 1

	return records


def AddDateToExcelFile(PubMed_Records):
	#add results to excel folder (including citations)...
	x=1
	filepath = GetDownloadPath() + "\\test_keyword_search.xlsx"
	try:
		book = load_workbook(filepath)
	except:
		print("creating excel file....")
		wb = openpyxl.Workbook()
		wb.save(filepath)
		book = load_workbook(filepath)
		sheet = book.active
		sheet['A1'] = "No."
		sheet['B1'] = "Title"
		sheet['C1'] = "Authors"
		sheet['D1'] = "Publication Date"
		sheet['E1'] = "DOI"
		sheet['F1'] = "Wiley Link"
		sheet['G1'] = "Citations"

	ws = book.worksheets[0]
	for record in PubMed_Records:
		for cell in ws["A"]:
			if cell.value is None:
				emptyRow = cell.row
				break
		else:
			emptyRow = cell.row + 1

		title = record.get("TI", "?")
		authors = ", ".join(record.get("AU", "?"))
		pub_date = record.get("DP", "?")[5:] + " " + record.get("DP", "?")[:-4]
		journal_name = record.get("JT", "?")
		doi = record.get("LID", "?")[:-6]
		wiley_link = "http://onlinelibrary.wiley.com/doi/full/" + record.get("LID", "?")[:-6]

		if citation_bool:
			excel_data = [str(x), title, authors, pub_date, doi, wiley_link, GetCitationNumber(wiley_link)]
		else:
			excel_data = [str(x), title, authors, pub_date, doi, wiley_link, "skipped"]
		
		x += 1

		for row in excel_data:
			ws.append(excel_data)
			break

		book.save(filepath)


def DoSearch(search_keyword):
	data = PubMedSearch(search_keyword)
	AddDateToExcelFile(data)

#process data without gui
#data = PubMedSearch(search_string)
#AddDateToExcelFile(data)



class Main:
	def __init__(self, parent):
		def AboutMe():
			#loads info
			CallDisplayAboutMe = DisplayAboutMe(parent)
			pass

		def StartApp():
			global display_message

			#CheckUpdates()
			menubar = tk.Menu(parent)
			filemenu = tk.Menu(menubar, tearoff=0)
			filemenu.add_command(label='Exit', command=parent.destroy)
			menubar.add_cascade(label='File', menu=filemenu)
			
			toolsmenu = tk.Menu(menubar, tearoff=0)
			menubar.add_cascade(label='Tools', menu=toolsmenu)

			helpmenu = tk.Menu(menubar, tearoff=0)
			helpmenu.add_command(label='About', command=AboutMe)
			menubar.add_cascade(label='Help', menu=helpmenu)
		
			parent.config(menu=menubar)

			rows = 0
			while rows < 50:
				parent.rowconfigure(rows, weight=1)
				parent.columnconfigure(rows, weight=1)
				rows += 1

			display_message = tk.StringVar() #message that shows user processing messages, error messages, etc
			display_message.set("Welcome to the JIAS PubMed Search Tool!")

			#Setup of processing/error message for a more user-friendly GUI
			main_info_display = tk.Label(parent, textvariable=display_message)
			main_info_display.grid(row=500, column=25)

			PubMed_Keyword_string = Entry(parent, width=30)
			PubMed_Keyword_string.grid(row=0, column=0)

			PubMed_Search_Button = Button(parent, text="Search PubMed and Create Excel Table with Results", command=lambda:DoSearch(PubMed_Keyword_string.get()))
			PubMed_Search_Button.grid(row=1, column=0)






			#begins the tkinter gui application
			pass
		StartApp()


def main():
	root = tk.Tk()
	root.title(__AppName__+' '+str(__version__))
	w=400; h=525
	sw = root.winfo_screenwidth()
	sh = root.winfo_screenheight()
	x = (sw - w) / 2
	y = (sh - h) / 2
	root.geometry('{0}x{1}+{2}+{3}'.format(w, h, int(x), int(y)))
	root.resizable(width=False, height=False)
	root.wm_iconbitmap('robot.ico')
	win = Main(root)
	root.mainloop()	


if __name__ == '__main__':
	main()